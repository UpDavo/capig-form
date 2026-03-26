from __future__ import annotations

import argparse
import math
import re
import unicodedata
import warnings
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_NEW_PATH = Path(r"c:\Users\Alexe\Downloads\base socios actualizacion DH.xlsx")
DEFAULT_CURRENT_PATH = PROJECT_ROOT / "data" / "CAPIG_DASHBOARD_HEIMDAL.xlsx"
DEFAULT_OUTPUT_PATH = PROJECT_ROOT / "data" / "SOCIOS_reemplazo_preparado.xlsx"

CURRENT_HEADERS = [
    "No. ",
    "RAZON_SOCIAL",
    "RUC",
    "FECHA_AFILIACION",
    "CIUDAD",
    "DIRECCION",
    "TELEFONO EMPRESA 1",
    "TELEFONO EMPRESA 2",
    "No. Colaboradores ",
    "EMAIL",
    "EMAIL2",
    "ACTIVIDAD",
    "NOMBRE REP_LEGAL",
    "CARGO",
    "NOMBRE GTE.  GRAL. ",
    "GÉNERO",
    "CONTACTO ",
    "Coreo Elsctronico",
    "NOMBRE GTE. FIN. ",
    "CONTACTO 2",
    "Coreo Elsctronico3",
    "NOMBRE GTE. RRHH ",
    "CONTACTO 4",
    "Coreo Elsctronico5",
    "GERENTE COMERC. ",
    "CONTACTO 6",
    "CORREO",
    "GERENTE PRODUCCCION",
    "CONTACTO 8",
    "Coreo Elsctronico9",
    "2019",
    "2020",
    "2021",
    "2022",
    "T2022",
    "2023",
    "T2023",
    "CRECIMIENTO",
    "SECTOR ",
    "TAMAÑO",
    "PORCENTAJE",
    "SEMAFORO",
    "ESTADO",
    "TOTAL POR TAMAÑO",
]

CURRENT_HEADERS_WITH_2024 = [
    "No. ",
    "RAZON_SOCIAL",
    "RUC",
    "FECHA_AFILIACION",
    "CIUDAD",
    "DIRECCION",
    "TELEFONO EMPRESA 1",
    "TELEFONO EMPRESA 2",
    "No. Colaboradores ",
    "EMAIL",
    "EMAIL2",
    "ACTIVIDAD",
    "NOMBRE REP_LEGAL",
    "CARGO",
    "NOMBRE GTE.  GRAL. ",
    "GÉNERO",
    "CONTACTO ",
    "Coreo Elsctronico",
    "NOMBRE GTE. FIN. ",
    "CONTACTO 2",
    "Coreo Elsctronico3",
    "NOMBRE GTE. RRHH ",
    "CONTACTO 4",
    "Coreo Elsctronico5",
    "GERENTE COMERC. ",
    "CONTACTO 6",
    "CORREO",
    "GERENTE PRODUCCCION",
    "CONTACTO 8",
    "Coreo Elsctronico9",
    "2019",
    "2020",
    "2021",
    "2022",
    "T2022",
    "2023",
    "2024",
    "T2023",
    "CRECIMIENTO",
    "SECTOR ",
    "TAMAÑO",
    "PORCENTAJE",
    "SEMAFORO",
    "ESTADO",
    "TOTAL POR TAMAÑO",
]

YEAR_MISSING_MARKERS = {
    "",
    "-",
    "N/A",
    "NA",
    "NO REPORTA",
    "NO REPORTO",
    "NO REPORTARON",
    "NO REPORTE",
    "NO REPORTA.",
    "NO REPORTO.",
    "NO INFO",
    "NO INFORMA",
    "SIN INFO",
    "SIN INFORMACION",
    "SIN INFORMACIÓN",
    "NONE",
    "NULL",
}

SIZE_LABELS = {
    1: "MICRO",
    2: "PEQUEÑA",
    3: "MEDIANA",
    4: "GRANDE",
}

MAPEO_ROWS = [
    ["Columna actual SOCIOS", "Fuente en archivo nuevo", "Accion aplicada"],
    ["No. ", "No. ", "Recalculada secuencialmente 1..N"],
    ["RAZON_SOCIAL", "RAZON_SOCIAL", "Copiada y limpiada"],
    ["RUC", "RUC", "Copiado como texto limpio; se conserva el formato visible del nuevo"],
    ["FECHA_AFILIACION", "FECHA_AFILIACION", "Copiada; se intenta mantenerla como fecha"],
    ["CIUDAD", "CIUDAD", "Copiada y limpiada"],
    ["DIRECCION", "DIRECCION", "Copiada y limpiada"],
    ["TELEFONO EMPRESA 1", "TELEFONO EMPRESA 1", "Copiada como texto"],
    ["TELEFONO EMPRESA 2", "TELEFONO EMPRESA 2", "Copiada como texto"],
    ["No. Colaboradores ", "No. Colaboradores ", "Copiada y normalizada"],
    ["EMAIL", "EMAIL", "Copiada y limpiada"],
    ["EMAIL2", "EMAIL2", "Copiada y limpiada"],
    ["ACTIVIDAD", "ACTIVIDAD", "Copiada y limpiada"],
    ["NOMBRE REP_LEGAL", "NOMBRE REP_LEGAL", "Copiada y limpiada"],
    ["CARGO", "CARGO", "Copiada y limpiada"],
    ["NOMBRE GTE.  GRAL. ", "NOMBRE GTE.  GRAL. ", "Copiada y limpiada"],
    ["GÉNERO", "(no existe en el nuevo)", "Recuperado del SOCIOS actual por RUC+RAZON_SOCIAL; si no coincide, queda vacio"],
    ["CONTACTO ", "CONTACTO ", "Copiada y desplazada a la posicion actual"],
    ["Coreo Elsctronico", "Coreo Elsctronico", "Copiada y desplazada a la posicion actual"],
    ["NOMBRE GTE. FIN. ", "NOMBRE GTE. FIN. ", "Copiada y desplazada a la posicion actual"],
    ["CONTACTO 2", "CONTACTO 2", "Copiada y desplazada a la posicion actual"],
    ["Coreo Elsctronico3", "Coreo Elsctronico3", "Copiada y desplazada a la posicion actual"],
    ["NOMBRE GTE. RRHH ", "NOMBRE GTE. RRHH ", "Copiada y desplazada a la posicion actual"],
    ["CONTACTO 4", "CONTACTO 4", "Copiada y desplazada a la posicion actual"],
    ["Coreo Elsctronico5", "Coreo Elsctronico5", "Copiada y desplazada a la posicion actual"],
    ["GERENTE COMERC. ", "GERENTE COMERC. ", "Copiada y desplazada a la posicion actual"],
    ["CONTACTO 6", "CONTACTO 6", "Copiada y desplazada a la posicion actual"],
    ["CORREO", "Coreo Elsctronico7", "Renombrada y desplazada a la posicion actual"],
    ["GERENTE PRODUCCCION", "GERENTE PRODUCCCION", "Copiada y desplazada a la posicion actual"],
    ["CONTACTO 8", "CONTACTO 8", "Copiada y desplazada a la posicion actual"],
    ["Coreo Elsctronico9", "Coreo Elsctronico9", "Copiada y desplazada a la posicion actual"],
    ["2019", "2019", "Normalizada como monto o 'no reporta'"],
    ["2020", "2020", "Normalizada como monto o 'no reporta'"],
    ["2021", "2021", "Normalizada como monto o 'no reporta'"],
    ["2022", "2022", "Normalizada como monto o 'no reporta'"],
    ["T2022", "T2022", "Recalculada con la formula original: si 2022 es 'no reporta', toma T2023; si no, clasifica 2022"],
    ["2023", "2023", "Normalizada como monto o 'no reporta'"],
    ["T2023", "T2023", "Recalculada con la formula original a partir de 2023"],
    ["CRECIMIENTO", "CRECIMIENTO", "Recalculado con la misma logica original entre T2022 y T2023"],
    ["SECTOR ", "SECTOR ", "Copiada y limpiada"],
    ["TAMAÑO", "TAMAÑO", "Recalculado a partir de 2023 para mantener compatibilidad con la hoja SOCIOS"],
    ["PORCENTAJE", "(no existe en el nuevo)", "Calculado como 1 / TOTAL POR TAMAÑO usando el TAMAÑO basado en 2023"],
    ["SEMAFORO", "SEMAFORO", "Recalculado con la logica original entre 2022 y 2023"],
    ["ESTADO", "ESTADO", "Copiado y limpiado"],
    ["TOTAL POR TAMAÑO", "Columna1", "No se usa el Columna1; se recalcula como conteo de filas del mismo tamaño basado en 2023"],
    ["(extra en nuevo)", "2024", "Se conserva como columna adicional; los derivados siguen anclados a 2022/2023 para no romper SOCIOS"],
]


def normalize_header(value: Any) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = re.sub(r"[^0-9A-Za-z]+", "_", text.upper()).strip("_")
    return text


def normalize_name(value: Any) -> str:
    text = clean_text(value).upper()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if math.isnan(value):
            return ""
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip()


def clean_display_ruc(value: Any) -> str:
    text = clean_text(value)
    text = text.replace('"', "").replace("'", "").strip()
    return re.sub(r"\s+", "", text)


def ruc_lookup_key(value: Any) -> str:
    return re.sub(r"\D", "", clean_text(value))


def is_year_missing(value: Any) -> bool:
    text = clean_text(value).upper()
    return text in YEAR_MISSING_MARKERS


def to_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)

    text = clean_text(value)
    if not text:
        return None

    compact = text.replace("$", "").replace(" ", "")
    if compact.upper() in YEAR_MISSING_MARKERS:
        return None

    has_comma = "," in compact
    has_dot = "." in compact

    if has_comma and has_dot:
        if compact.rfind(".") > compact.rfind(","):
            compact = compact.replace(",", "")
        else:
            compact = compact.replace(".", "").replace(",", ".")
    elif has_comma:
        if re.fullmatch(r"-?\d{1,3}(,\d{3})+(\.\d+)?", compact):
            compact = compact.replace(",", "")
        else:
            parts = compact.split(",")
            if len(parts) == 2 and len(parts[1]) <= 2:
                compact = f"{parts[0]}.{parts[1]}"
            else:
                compact = compact.replace(",", "")
    elif compact.count(".") > 1:
        compact = compact.replace(".", "")

    compact = re.sub(r"[^0-9\.-]", "", compact)
    if compact in {"", "-", ".", "-."}:
        return None
    try:
        return float(compact)
    except ValueError:
        return None


def normalize_year_value(value: Any) -> str | int | float:
    if is_year_missing(value):
        return "no reporta"
    number = to_number(value)
    if number is None:
        return "no reporta"
    if number.is_integer():
        return int(number)
    return round(number, 2)


def classify_size_code(value: str | int | float) -> int:
    number = to_number(value)
    if number is None:
        return 1
    if number >= 5_000_000:
        return 4
    if number >= 1_000_000:
        return 3
    if number >= 100_000:
        return 2
    return 1


def calc_t2023(value_2023: str | int | float) -> int:
    if is_year_missing(value_2023):
        return 1
    return classify_size_code(value_2023)


def calc_t2022(value_2022: str | int | float, t2023: int) -> int:
    if is_year_missing(value_2022):
        return t2023
    return classify_size_code(value_2022)


def calc_crecimiento(t2022: int, t2023: int) -> str:
    if t2022 == t2023:
        return "IGUAL"
    if t2022 > t2023:
        return "DECRECIÓ"
    return "CRECIÓ"


def calc_tamano(value_2023: str | int | float) -> str:
    if is_year_missing(value_2023):
        return "MICRO"
    return SIZE_LABELS[classify_size_code(value_2023)]


def excel_like_kind(value: Any) -> Tuple[str, float | str]:
    if is_year_missing(value):
        return ("text", "NO REPORTA")
    number = to_number(value)
    if number is not None:
        return ("number", number)
    return ("text", clean_text(value).upper())


def excel_like_compare(left: Any, right: Any) -> int:
    kind_left, value_left = excel_like_kind(left)
    kind_right, value_right = excel_like_kind(right)

    if kind_left == kind_right == "number":
        return (value_left > value_right) - (value_left < value_right)
    if kind_left == kind_right == "text":
        return (value_left > value_right) - (value_left < value_right)
    if kind_left == "text" and kind_right == "number":
        return 1
    return -1


def calc_semaforo(value_2022: Any, value_2023: Any) -> str:
    comparison = excel_like_compare(value_2023, value_2022)
    if comparison > 0:
        return "VERDE"
    if comparison < 0:
        return "ROJO"
    return "AMARILLO"


def latest_reported_value(year_values: Dict[int, Any]) -> Tuple[int | None, Any]:
    reported = [(year, value) for year, value in sorted(year_values.items()) if not is_year_missing(value)]
    if not reported:
        return (None, "no reporta")
    return reported[-1]


def latest_reported_pair(year_values: Dict[int, Any]) -> Tuple[Tuple[int | None, Any], Tuple[int | None, Any]]:
    reported = [(year, value) for year, value in sorted(year_values.items()) if not is_year_missing(value)]
    if not reported:
        empty = (None, "no reporta")
        return empty, empty
    if len(reported) == 1:
        return reported[0], reported[0]
    return reported[-2], reported[-1]


def calc_current_tamano(year_values: Dict[int, Any]) -> str:
    _, latest_value = latest_reported_value(year_values)
    return calc_tamano(latest_value)


def calc_current_crecimiento(year_values: Dict[int, Any]) -> str:
    (_, previous_value), (_, current_value) = latest_reported_pair(year_values)
    return calc_crecimiento(classify_size_code(previous_value), classify_size_code(current_value))


def calc_current_semaforo(year_values: Dict[int, Any]) -> str:
    (_, previous_value), (_, current_value) = latest_reported_pair(year_values)
    return calc_semaforo(previous_value, current_value)


def clean_generic_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    if isinstance(value, float):
        if math.isnan(value):
            return ""
        if value.is_integer():
            return str(int(value))
        return str(value)
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def normalize_estado(value: Any) -> str:
    return clean_text(value).upper()


def find_header_row(worksheet, required_headers: Iterable[str], max_scan_rows: int = 20) -> int:
    expected = {normalize_header(item) for item in required_headers}
    max_rows = worksheet.max_row or max_scan_rows
    max_cols = min(worksheet.max_column or 200, 200)
    for row_idx in range(1, min(max_rows, max_scan_rows) + 1):
        normalized = {normalize_header(worksheet.cell(row_idx, col).value) for col in range(1, max_cols + 1)}
        if expected.issubset(normalized):
            return row_idx
    raise RuntimeError(f"No se encontro fila de encabezados con: {sorted(expected)}")


def load_rows(worksheet, header_row_idx: int) -> Tuple[List[str], List[Tuple[int, Dict[str, Any]]]]:
    max_cols = min(worksheet.max_column or 200, 200)
    max_rows = worksheet.max_row or header_row_idx
    header_values = [worksheet.cell(header_row_idx, col).value for col in range(1, max_cols + 1)]
    used_columns = [idx for idx, value in enumerate(header_values, start=1) if value not in (None, "")]
    if not used_columns:
        raise RuntimeError("La fila de encabezados no contiene columnas utilizables.")

    last_col = max(used_columns)
    header_values = header_values[:last_col]
    normalized_headers = [normalize_header(value) for value in header_values]

    rows: List[Tuple[int, Dict[str, Any]]] = []
    for row_idx in range(header_row_idx + 1, max_rows + 1):
        values = [worksheet.cell(row_idx, col).value for col in range(1, last_col + 1)]
        if not any(value not in (None, "") for value in values):
            continue
        row_dict = {normalized_headers[idx]: values[idx] for idx in range(len(normalized_headers))}
        rows.append((row_idx, row_dict))
    return [clean_text(value) for value in header_values], rows


def old_gender_maps(current_path: Path) -> Tuple[Dict[Tuple[str, str], str], Dict[str, str], Dict[str, set[str]], Dict[str, set[str]]]:
    workbook = load_workbook(current_path, data_only=True, read_only=False)
    worksheet = workbook["SOCIOS"]
    header_row_idx = find_header_row(worksheet, ["RUC", "RAZON_SOCIAL"])
    _, rows = load_rows(worksheet, header_row_idx)

    by_exact: Dict[Tuple[str, str], str] = {}
    by_ruc_candidates: Dict[str, set[str]] = defaultdict(set)
    reasons_by_ruc: Dict[str, set[str]] = defaultdict(set)

    for _, row in rows:
        ruc_display = clean_display_ruc(row.get("RUC"))
        ruc_key = ruc_lookup_key(ruc_display)
        razon = clean_text(row.get("RAZON_SOCIAL"))
        razon_key = normalize_name(razon)
        genero = clean_text(row.get("GENERO"))
        if ruc_key and razon:
            reasons_by_ruc[ruc_key].add(razon)
        if ruc_key and razon_key and genero:
            by_exact[(ruc_key, razon_key)] = genero
            by_ruc_candidates[ruc_key].add(genero)

    by_ruc_unique = {ruc: next(iter(genders)) for ruc, genders in by_ruc_candidates.items() if len(genders) == 1}
    return by_exact, by_ruc_unique, by_ruc_candidates, reasons_by_ruc


def resolve_genero(ruc_key: str, razon_key: str, exact_map: Dict[Tuple[str, str], str], by_ruc_unique: Dict[str, str]) -> str:
    if (ruc_key, razon_key) in exact_map:
        return exact_map[(ruc_key, razon_key)]
    return by_ruc_unique.get(ruc_key, "")


def set_column_widths(worksheet) -> None:
    widths = defaultdict(int)
    for row in worksheet.iter_rows():
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            widths[cell.column] = min(max(widths[cell.column], len(value) + 2), 60)
    for column_idx, width in widths.items():
        worksheet.column_dimensions[get_column_letter(column_idx)].width = width


def style_table_sheet(worksheet) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    set_column_widths(worksheet)


def apply_replacement_formats(worksheet) -> None:
    style_table_sheet(worksheet)
    for row in range(2, worksheet.max_row + 1):
        worksheet.cell(row, 1).number_format = "0"
        worksheet.cell(row, 3).number_format = "@"
        worksheet.cell(row, 4).number_format = "dd/mm/yyyy"
        worksheet.cell(row, 35).number_format = "0"
        worksheet.cell(row, 37).number_format = "0"
        worksheet.cell(row, 41).number_format = "0.00%"
        worksheet.cell(row, 44).number_format = "0"
        for col in (31, 32, 33, 34, 36):
            if isinstance(worksheet.cell(row, col).value, (int, float)):
                worksheet.cell(row, col).number_format = "#,##0.00"


def apply_new_sheet_formats(worksheet) -> None:
    style_table_sheet(worksheet)
    for row in range(2, worksheet.max_row + 1):
        worksheet.cell(row, 1).number_format = "0"
        worksheet.cell(row, 3).number_format = "@"
        worksheet.cell(row, 4).number_format = "dd/mm/yyyy"
        worksheet.cell(row, 34).number_format = "0"
        worksheet.cell(row, 37).number_format = "0"
        for col in (30, 31, 32, 33, 35, 36):
            if isinstance(worksheet.cell(row, col).value, (int, float)):
                worksheet.cell(row, col).number_format = "#,##0.00"


def apply_replacement_2024_formats(worksheet) -> None:
    style_table_sheet(worksheet)
    for row in range(2, worksheet.max_row + 1):
        worksheet.cell(row, 1).number_format = "0"
        worksheet.cell(row, 3).number_format = "@"
        worksheet.cell(row, 4).number_format = "dd/mm/yyyy"
        worksheet.cell(row, 35).number_format = "0"
        worksheet.cell(row, 38).number_format = "0"
        worksheet.cell(row, 42).number_format = "0.00%"
        worksheet.cell(row, 45).number_format = "0"
        for col in (31, 32, 33, 34, 36, 37):
            if isinstance(worksheet.cell(row, col).value, (int, float)):
                worksheet.cell(row, col).number_format = "#,##0.00"


def build_output(new_path: Path, current_path: Path, output_path: Path) -> Dict[str, Any]:
    current_exact_gender, current_unique_gender, current_gender_candidates, old_reasons_by_ruc = old_gender_maps(current_path)

    new_wb = load_workbook(new_path, data_only=False, read_only=False)
    new_ws = new_wb["Hoja1"]
    new_header_row = find_header_row(new_ws, ["RUC", "RAZON_SOCIAL", "FECHA_AFILIACION"])
    new_headers, new_rows = load_rows(new_ws, new_header_row)

    clean_new_records: List[Dict[str, Any]] = []
    replacement_records: List[Dict[str, Any]] = []
    incidencias: List[List[Any]] = [["TIPO", "RUC", "RAZON_SOCIAL", "DETALLE"]]

    duplicates_new: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    new_reasons_by_ruc: Dict[str, set[str]] = defaultdict(set)

    for visible_index, (row_idx, row) in enumerate(new_rows, start=1):
        razon_social = clean_text(row.get("RAZON_SOCIAL"))
        ruc_display = clean_display_ruc(row.get("RUC"))
        ruc_key = ruc_lookup_key(ruc_display)
        razon_key = normalize_name(razon_social)

        if not razon_social and not ruc_display:
            continue

        value_2019 = normalize_year_value(row.get("2019"))
        value_2020 = normalize_year_value(row.get("2020"))
        value_2021 = normalize_year_value(row.get("2021"))
        value_2022 = normalize_year_value(row.get("2022"))
        value_2023 = normalize_year_value(row.get("2023"))
        value_2024 = normalize_year_value(row.get("2024"))
        year_values = {
            2019: value_2019,
            2020: value_2020,
            2021: value_2021,
            2022: value_2022,
            2023: value_2023,
            2024: value_2024,
        }

        t2023 = calc_t2023(value_2023)
        t2022 = calc_t2022(value_2022, t2023)
        crecimiento = calc_crecimiento(t2022, t2023)
        tamano = calc_tamano(value_2023)
        semaforo = calc_semaforo(value_2022, value_2023)
        current_tamano = calc_current_tamano(year_values)
        current_crecimiento = calc_current_crecimiento(year_values)
        current_semaforo = calc_current_semaforo(year_values)
        genero = resolve_genero(ruc_key, razon_key, current_exact_gender, current_unique_gender)

        new_record = {
            "No. ": visible_index,
            "RAZON_SOCIAL": razon_social,
            "RUC": ruc_display,
            "FECHA_AFILIACION": row.get("FECHA_AFILIACION"),
            "CIUDAD": clean_generic_value(row.get("CIUDAD")),
            "DIRECCION": clean_generic_value(row.get("DIRECCION")),
            "TELEFONO EMPRESA 1": clean_generic_value(row.get("TELEFONO_EMPRESA_1")),
            "TELEFONO EMPRESA 2": clean_generic_value(row.get("TELEFONO_EMPRESA_2")),
            "No. Colaboradores ": clean_generic_value(row.get("NO_COLABORADORES")),
            "EMAIL": clean_generic_value(row.get("EMAIL")),
            "EMAIL2": clean_generic_value(row.get("EMAIL2")),
            "ACTIVIDAD": clean_generic_value(row.get("ACTIVIDAD")),
            "NOMBRE REP_LEGAL": clean_generic_value(row.get("NOMBRE_REP_LEGAL")),
            "CARGO": clean_generic_value(row.get("CARGO")),
            "NOMBRE GTE.  GRAL. ": clean_generic_value(row.get("NOMBRE_GTE_GRAL")),
            "GÉNERO": genero,
            "CONTACTO ": clean_generic_value(row.get("CONTACTO")),
            "Coreo Elsctronico": clean_generic_value(row.get("COREO_ELSCTRONICO")),
            "NOMBRE GTE. FIN. ": clean_generic_value(row.get("NOMBRE_GTE_FIN")),
            "CONTACTO 2": clean_generic_value(row.get("CONTACTO_2")),
            "Coreo Elsctronico3": clean_generic_value(row.get("COREO_ELSCTRONICO3")),
            "NOMBRE GTE. RRHH ": clean_generic_value(row.get("NOMBRE_GTE_RRHH")),
            "CONTACTO 4": clean_generic_value(row.get("CONTACTO_4")),
            "Coreo Elsctronico5": clean_generic_value(row.get("COREO_ELSCTRONICO5")),
            "GERENTE COMERC. ": clean_generic_value(row.get("GERENTE_COMERC")),
            "CONTACTO 6": clean_generic_value(row.get("CONTACTO_6")),
            "CORREO": clean_generic_value(row.get("COREO_ELSCTRONICO7")),
            "GERENTE PRODUCCCION": clean_generic_value(row.get("GERENTE_PRODUCCCION")),
            "CONTACTO 8": clean_generic_value(row.get("CONTACTO_8")),
            "Coreo Elsctronico9": clean_generic_value(row.get("COREO_ELSCTRONICO9")),
            "2019": value_2019,
            "2020": value_2020,
            "2021": value_2021,
            "2022": value_2022,
            "T2022": t2022,
            "2023": value_2023,
            "T2023": t2023,
            "CRECIMIENTO": crecimiento,
            "SECTOR ": clean_generic_value(row.get("SECTOR")),
            "TAMAÑO": tamano,
            "PORCENTAJE": None,
            "SEMAFORO": semaforo,
            "ESTADO": normalize_estado(row.get("ESTADO")),
            "TOTAL POR TAMAÑO": None,
            "_TAMAÑO_ACTUAL": current_tamano,
            "_PORCENTAJE_ACTUAL": None,
            "_SEMAFORO_ACTUAL": current_semaforo,
            "_CRECIMIENTO_ACTUAL": current_crecimiento,
            "_TOTAL_POR_TAMAÑO_ACTUAL": None,
            "_source_row": row_idx,
            "_ruc_key": ruc_key,
            "_razon_key": razon_key,
            "_2024": value_2024,
            "_columna1_original": clean_generic_value(row.get("COLUMNA1")),
        }

        clean_new_records.append(new_record)
        replacement_records.append(new_record)
        if ruc_key:
            new_reasons_by_ruc[ruc_key].add(razon_social)
            duplicates_new[ruc_key].append(new_record)

        if not genero:
            incidencias.append(
                [
                    "SIN_GENERO_RECUPERABLE",
                    ruc_display,
                    razon_social,
                    f"Fila nueva {row_idx}. No hubo coincidencia segura contra el SOCIOS actual.",
                ]
            )

    size_counter = Counter(record["TAMAÑO"] for record in replacement_records if record["TAMAÑO"])
    for record in replacement_records:
        size_total = size_counter.get(record["TAMAÑO"], 0)
        record["TOTAL POR TAMAÑO"] = size_total if size_total else ""
        record["PORCENTAJE"] = (1 / size_total) if size_total else ""

    size_counter_current = Counter(record["_TAMAÑO_ACTUAL"] for record in replacement_records if record["_TAMAÑO_ACTUAL"])
    for record in replacement_records:
        size_total_current = size_counter_current.get(record["_TAMAÑO_ACTUAL"], 0)
        record["_TOTAL_POR_TAMAÑO_ACTUAL"] = size_total_current if size_total_current else ""
        record["_PORCENTAJE_ACTUAL"] = (1 / size_total_current) if size_total_current else ""

    for ruc_key, items in duplicates_new.items():
        if not ruc_key or len(items) == 1:
            continue
        detail = "; ".join(f"fila {item['_source_row']}: {item['RAZON_SOCIAL']}" for item in items)
        incidencias.append(["RUC_DUPLICADO_EN_NUEVA", items[0]["RUC"], "", detail])

    old_ruc_set = set(old_reasons_by_ruc)
    new_ruc_set = set(new_reasons_by_ruc)
    only_new = sorted(new_ruc_set - old_ruc_set)
    only_old = sorted(old_ruc_set - new_ruc_set)

    for ruc_key in only_new:
        detalle = "; ".join(sorted(new_reasons_by_ruc[ruc_key]))
        incidencias.append(["RUC_SOLO_EN_NUEVA", ruc_key, "", detalle])

    for ruc_key in only_old:
        detalle = "; ".join(sorted(old_reasons_by_ruc[ruc_key]))
        incidencias.append(["RUC_SOLO_EN_ACTUAL", ruc_key, "", detalle])

    summary_rows = [
        ["METRICA", "VALOR"],
        ["Filas utiles del archivo nuevo", len(replacement_records)],
        ["RUC unicos en archivo nuevo", len(new_ruc_set)],
        ["RUC duplicados en archivo nuevo", sum(1 for items in duplicates_new.values() if len(items) > 1)],
        ["Filas sin genero recuperable", sum(1 for record in replacement_records if not record["GÉNERO"])],
        ["RUC solo en nueva", len(only_new)],
        ["RUC solo en actual", len(only_old)],
        ["Archivo nuevo analizado", str(new_path)],
        ["Archivo actual usado como guia", str(current_path)],
        ["Salida generada", str(output_path)],
    ]

    output_wb = Workbook()
    ws_replacement = output_wb.active
    ws_replacement.title = "SOCIOS_REEMPLAZO"
    ws_replacement.append(CURRENT_HEADERS)
    for record in replacement_records:
        ws_replacement.append([record[header] for header in CURRENT_HEADERS])
    apply_replacement_formats(ws_replacement)

    ws_replacement_2024 = output_wb.create_sheet("SOCIOS_REEMPLAZO_2024")
    ws_replacement_2024.append(CURRENT_HEADERS_WITH_2024)
    for record in replacement_records:
        ws_replacement_2024.append(
            [
                record["No. "],
                record["RAZON_SOCIAL"],
                record["RUC"],
                record["FECHA_AFILIACION"],
                record["CIUDAD"],
                record["DIRECCION"],
                record["TELEFONO EMPRESA 1"],
                record["TELEFONO EMPRESA 2"],
                record["No. Colaboradores "],
                record["EMAIL"],
                record["EMAIL2"],
                record["ACTIVIDAD"],
                record["NOMBRE REP_LEGAL"],
                record["CARGO"],
                record["NOMBRE GTE.  GRAL. "],
                record["GÉNERO"],
                record["CONTACTO "],
                record["Coreo Elsctronico"],
                record["NOMBRE GTE. FIN. "],
                record["CONTACTO 2"],
                record["Coreo Elsctronico3"],
                record["NOMBRE GTE. RRHH "],
                record["CONTACTO 4"],
                record["Coreo Elsctronico5"],
                record["GERENTE COMERC. "],
                record["CONTACTO 6"],
                record["CORREO"],
                record["GERENTE PRODUCCCION"],
                record["CONTACTO 8"],
                record["Coreo Elsctronico9"],
                record["2019"],
                record["2020"],
                record["2021"],
                record["2022"],
                record["T2022"],
                record["2023"],
                record["_2024"],
                record["T2023"],
                record["CRECIMIENTO"],
                record["SECTOR "],
                record["TAMAÑO"],
                record["PORCENTAJE"],
                record["SEMAFORO"],
                record["ESTADO"],
                record["TOTAL POR TAMAÑO"],
            ]
        )
    apply_replacement_2024_formats(ws_replacement_2024)

    ws_new = output_wb.create_sheet("NUEVA_LIMPIA")
    new_headers_out = [
        "No. ",
        "RAZON_SOCIAL",
        "RUC",
        "FECHA_AFILIACION",
        "CIUDAD",
        "DIRECCION",
        "TELEFONO EMPRESA 1",
        "TELEFONO EMPRESA 2",
        "No. Colaboradores ",
        "EMAIL",
        "EMAIL2",
        "ACTIVIDAD",
        "NOMBRE REP_LEGAL",
        "CARGO",
        "NOMBRE GTE.  GRAL. ",
        "CONTACTO ",
        "Coreo Elsctronico",
        "NOMBRE GTE. FIN. ",
        "CONTACTO 2",
        "Coreo Elsctronico3",
        "NOMBRE GTE. RRHH ",
        "CONTACTO 4",
        "Coreo Elsctronico5",
        "GERENTE COMERC. ",
        "CONTACTO 6",
        "Coreo Elsctronico7",
        "GERENTE PRODUCCCION",
        "CONTACTO 8",
        "Coreo Elsctronico9",
        "2019",
        "2020",
        "2021",
        "2022",
        "T2022",
        "2023",
        "2024",
        "T2023",
        "CRECIMIENTO",
        "SECTOR ",
        "TAMAÑO",
        "SEMAFORO",
        "ESTADO",
        "Columna1",
    ]
    ws_new.append(new_headers_out)
    for record in clean_new_records:
        ws_new.append(
            [
                record["No. "],
                record["RAZON_SOCIAL"],
                record["RUC"],
                record["FECHA_AFILIACION"],
                record["CIUDAD"],
                record["DIRECCION"],
                record["TELEFONO EMPRESA 1"],
                record["TELEFONO EMPRESA 2"],
                record["No. Colaboradores "],
                record["EMAIL"],
                record["EMAIL2"],
                record["ACTIVIDAD"],
                record["NOMBRE REP_LEGAL"],
                record["CARGO"],
                record["NOMBRE GTE.  GRAL. "],
                record["CONTACTO "],
                record["Coreo Elsctronico"],
                record["NOMBRE GTE. FIN. "],
                record["CONTACTO 2"],
                record["Coreo Elsctronico3"],
                record["NOMBRE GTE. RRHH "],
                record["CONTACTO 4"],
                record["Coreo Elsctronico5"],
                record["GERENTE COMERC. "],
                record["CONTACTO 6"],
                record["CORREO"],
                record["GERENTE PRODUCCCION"],
                record["CONTACTO 8"],
                record["Coreo Elsctronico9"],
                record["2019"],
                record["2020"],
                record["2021"],
                record["2022"],
                record["T2022"],
                record["2023"],
                record["_2024"],
                record["T2023"],
                record["CRECIMIENTO"],
                record["SECTOR "],
                record["TAMAÑO"],
                record["SEMAFORO"],
                record["ESTADO"],
                record["_columna1_original"],
            ]
        )
    apply_new_sheet_formats(ws_new)

    ws_mapping = output_wb.create_sheet("MAPEO")
    for row in MAPEO_ROWS:
        ws_mapping.append(row)
    style_table_sheet(ws_mapping)

    ws_summary = output_wb.create_sheet("RESUMEN")
    for row in summary_rows:
        ws_summary.append(row)
    style_table_sheet(ws_summary)

    ws_issues = output_wb.create_sheet("INCIDENCIAS")
    for row in incidencias:
        ws_issues.append(row)
    style_table_sheet(ws_issues)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_wb.save(output_path)

    return {
        "output_path": output_path,
        "rows": len(replacement_records),
        "unique_new_ruc": len(new_ruc_set),
        "duplicate_new_ruc": sum(1 for items in duplicates_new.values() if len(items) > 1),
        "missing_gender_rows": sum(1 for record in replacement_records if not record["GÉNERO"]),
        "only_new": len(only_new),
        "only_old": len(only_old),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Prepara una hoja SOCIOS lista para reemplazo manual.")
    parser.add_argument("--new", default=str(DEFAULT_NEW_PATH), help="Ruta del archivo nuevo que vas a reemplazar.")
    parser.add_argument("--current", default=str(DEFAULT_CURRENT_PATH), help="Ruta del archivo actual usado como guia.")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT_PATH), help="Ruta del archivo de salida.")
    args = parser.parse_args()

    result = build_output(Path(args.new), Path(args.current), Path(args.output))
    print("Archivo generado:", result["output_path"])
    print("Filas preparadas:", result["rows"])
    print("RUC unicos en nueva:", result["unique_new_ruc"])
    print("RUC duplicados en nueva:", result["duplicate_new_ruc"])
    print("Filas sin genero recuperable:", result["missing_gender_rows"])
    print("RUC solo en nueva:", result["only_new"])
    print("RUC solo en actual:", result["only_old"])


if __name__ == "__main__":
    main()
